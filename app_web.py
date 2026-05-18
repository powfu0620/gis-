import os
import json
import re
import hashlib
import streamlit as st
import geopandas as gpd
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import JsonOutputParser
from langchain_openai import ChatOpenAI

# ==========================================
# 1. 页面基本配置
# ==========================================
st.set_page_config(
    page_title="GIS AI 属性表交互排版系统",
    page_icon="🌍",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 阿里云百炼大模型 API 配置
os.environ["DASHSCOPE_API_KEY"] = "sk-ea10027b5be245e0995c655488382266"  # 💥 请在此处替换为您真实的 API Key
BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
MODEL_NAME = "qwen-plus"

USER_DB_FILE = "users_db_web.json"


# ==========================================
# 2. 核心后端逻辑
# ==========================================
def apply_styles_to_excel(excel_path, style_config):
    """【终极智能容错版】Excel 样式安全渲染"""
    wb = load_workbook(excel_path)
    ws = wb.active
    max_row, max_col = ws.max_row, ws.max_column
    hex_color_pattern = re.compile(r'^[0-9a-fA-F]{6}$|^[0-9a-fA-F]{8}$')

    for target in style_config.get("targets", []):
        range_type = target.get("range_type")
        specs = target.get("styles", {})

        font_kwargs = {}
        if "font_name" in specs and specs["font_name"]: font_kwargs["name"] = str(specs["font_name"]).strip()
        if "font_size" in specs and specs["font_size"]:
            size_digits = re.findall(r'\d+', str(specs["font_size"]))
            if size_digits: font_kwargs["size"] = int(size_digits[0])

        # ---- 文字颜色安全清洗 ----
        if "font_color" in specs and specs["font_color"]:
            color_str = str(specs["font_color"]).strip().lower()
            # 常用颜色单词映射
            color_map = {
                "blue": "0000FF", "yellow": "FFFF00", "red": "FF0000",
                "green": "008000", "gray": "808080", "black": "000000", "white": "FFFFFF"
            }
            if color_str in color_map:
                clean_color = color_map[color_str]
            else:
                clean_color = color_str.replace("#", "")

            if hex_color_pattern.match(clean_color):
                font_kwargs["color"] = clean_color

        if "bold" in specs: font_kwargs["bold"] = bool(specs["bold"])
        if "italic" in specs: font_kwargs["italic"] = bool(specs["italic"])
        font_obj = Font(**font_kwargs) if font_kwargs else None

        # ---- 💥 背景颜色安全清洗（加入常用英文单词转换，解决不染色问题） 💥 ----
        fill_obj = None
        if "bg_color" in specs and specs["bg_color"]:
            color_str = str(specs["bg_color"]).strip().lower()

            # 转换成适合 Excel 排版的淡雅/标准商用 hex 颜色
            color_map = {
                "blue": "ADD8E6",  # 浅蓝
                "yellow": "FFFFE0",  # 浅黄
                "red": "FFC0CB",  # 浅红（粉）
                "green": "E0EEEE",  # 浅绿
                "gray": "F5F5F5"  # 浅灰
            }
            if color_str in color_map:
                clean_bg = color_map[color_str]
            else:
                clean_bg = color_str.replace("#", "")

            if hex_color_pattern.match(clean_bg):
                fill_obj = PatternFill(start_color=clean_bg, end_color=clean_bg, fill_type="solid")

        # ---- 对齐处理 ----
        align_obj = None
        if "alignment" in specs and specs["alignment"]:
            align_mapping = {"居中": "center", "左对齐": "left", "右对齐": "right", "center": "center", "left": "left",
                             "right": "right"}
            align_obj = Alignment(horizontal=align_mapping.get(specs["alignment"], "center"), vertical="center")

        # ---- 💥 确定影响范围（加入模糊容错匹配，彻底解决大模型乱吐 range_type 的问题） 💥 ----
        cells_to_modify = []
        range_str = str(range_type).lower()

        # 只要 AI 吐出的词包含 "header" 或 "head"，就认定是表头
        if "header" in range_str or "head" in range_str:
            cells_to_modify = [ws.cell(row=1, column=c) for c in range(1, max_col + 1)]

        # 只要 AI 吐出的词包含 "data" 或 "body" 或 "cell"，就认定是数据行
        elif "data" in range_str or "body" in range_str or "cell" in range_str:
            cells_to_modify = [ws.cell(row=r, column=c) for r in range(2, max_row + 1) for c in range(1, max_col + 1)]

        # 只要 AI 吐出的词包含 "all" 或 "table"，就全表修改
        elif "all" in range_str or "table" in range_str:
            cells_to_modify = [ws.cell(row=r, column=c) for r in range(1, max_row + 1) for c in range(1, max_col + 1)]

        elif "column" in range_str or "col" in range_str:
            col_name = target.get("column_name", "")
            col_idx = None
            for c in range(1, max_col + 1):
                if ws.cell(row=1, column=c).value == col_name:
                    col_idx = c
                    break
            if col_idx:
                cells_to_modify = [ws.cell(row=r, column=col_idx) for r in range(2, max_row + 1)]

        # 批量应用样式
        for cell in cells_to_modify:
            if font_obj:
                cur = cell.font
                cell.font = Font(name=font_kwargs.get("name", cur.name if cur else "宋体"),
                                 size=font_kwargs.get("size", cur.size if cur else 11),
                                 color=font_kwargs.get("color", cur.color if cur else "000000"),
                                 bold=font_kwargs.get("bold", cur.bold if cur else False))
            if fill_obj: cell.fill = fill_obj
            if align_obj: cell.alignment = align_obj

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 5, 12)
    wb.save(excel_path)
    wb.close()


def ask_ai_to_parse_instruction(user_instruction):
    """LangChain 大模型指令解析"""
    llm = ChatOpenAI(api_key=os.environ.get("DASHSCOPE_API_KEY"), base_url=BASE_URL, model=MODEL_NAME, temperature=0.0)
    system_prompt = (
        "你是一个精通 Excel 样式解析的 AI 助手。你的任务是将用户的中文指令转化为精准的 JSON 格式。\n"
        "【严格按照以下规则输出，不要包含 markdown 标记代码块包裹】:\n"
        "1. targets 为数组。\n"
        "2. range_type 的可选值必须严格控制在: 'header', 'data', 'all', 'column' 之中。\n"
        "3. styles 内部包含: font_name, font_size, font_color, bg_color, bold, alignment。\n"
        "4. 颜色请尽量直接输出 6 位 16 进制代码（不要带#号）。如果确实无法把握，再输出标准英文颜色单词（如 red, blue）。"
    )
    prompt_template = ChatPromptTemplate.from_messages(
        [("system", system_prompt), ("user", "用户的排版修改指令是：{instruction}")])
    return (prompt_template | llm | JsonOutputParser()).invoke({"instruction": user_instruction})


# ==========================================
# 3. 会员数据库与验证管理
# ==========================================
def load_users():
    if not os.path.exists(USER_DB_FILE): return {}
    with open(USER_DB_FILE, "r", encoding="utf-8") as f: return json.load(f)


def save_users(users):
    with open(USER_DB_FILE, "w", encoding="utf-8") as f: json.dump(users, f, ensure_ascii=False, indent=4)


if "logged_in" not in st.session_state: st.session_state.logged_in = False
if "current_user" not in st.session_state: st.session_state.current_user = None
if "pay_unlocked" not in st.session_state: st.session_state.pay_unlocked = False

# ==========================================
# 4. 网页前端布局
# ==========================================
st.title("🌍 GIS 属性表交互式 AI 智能排版系统")
st.markdown("##### 结合 **LangChain 大模型** 与 **空间数据管理**，让属性表样式调整像聊天一样简单。")
st.write("---")

# 侧边栏
with st.sidebar:
    st.header("🔑 商业授权与用户中心")
    if st.session_state.logged_in:
        st.success(f"🟢 已登录会员: {st.session_state.current_user}")
        if st.button("注销账户"):
            st.session_state.logged_in = False
            st.session_state.current_user = None
            st.rerun()
    else:
        auth_tab1, auth_tab2 = st.tabs(["会员登录", "QQ邮箱注册"])
        users = load_users()
        with auth_tab1:
            login_email = st.text_input("QQ 邮箱", key="login_email")
            login_pwd = st.text_input("密码", type="password", key="login_pwd")
            if st.button("立即登录", use_container_width=True):
                hashed_p = hashlib.sha256(login_pwd.encode()).hexdigest()
                if login_email in users and users[login_email] == hashed_p:
                    st.session_state.logged_in = True
                    st.session_state.current_user = login_email
                    st.rerun()
                else:
                    st.error("邮箱或密码不正确！")
        with auth_tab2:
            reg_email = st.text_input("QQ 邮箱", key="reg_email")
            reg_pwd = st.text_input("设置密码", type="password", key="reg_pwd")
            if st.button("提交注册", use_container_width=True):
                if not re.match(r'^[1-9][0-9]{4,10}@qq\.com$', reg_email):
                    st.error("请输入规范的 QQ 邮箱！")
                elif len(reg_pwd) < 6:
                    st.error("密码不能少于 6 位！")
                elif reg_email in users:
                    st.warning("该邮箱已被注册！")
                else:
                    users[reg_email] = hashlib.sha256(reg_pwd.encode()).hexdigest()
                    save_users(users)
                    st.success("注册成功！请切换到登录页。")

    st.write("---")
    st.markdown("<p style='text-align: center; color: gray;'>💡 支持免登录单次解锁：</p>", unsafe_allow_html=True)
    if st.session_state.pay_unlocked:
        st.info("🟢 微信支付成功，单次权限已解锁！")
        if st.button("清除单次授权"):
            st.session_state.pay_unlocked = False
            st.rerun()
    else:
        with st.popover("🟢 微信扫码快捷支付 (￥1.00)", use_container_width=True):
            if os.path.exists("wechat_pay.png"):
                st.image("wechat_pay.png", width=200)
            else:
                st.warning("请上传 wechat_pay.png")
            if st.button("我已支付，立即解锁", type="primary", use_container_width=True):
                st.session_state.pay_unlocked = True
                st.rerun()

# 主界面区
col_left, col_right = st.columns([1, 1])

with col_left:
    st.subheader("🛠️ 1. 动态输入数据源")
    uploaded_files = st.file_uploader(
        "请选择并上传您的矢量图层文件（GIS数据请同时拖入 .shp, .shx, .dbf 三个文件）",
        type=["shp", "shx", "dbf"],
        accept_multiple_files=True
    )

    st.write("")
    st.subheader("🤖 2. 输入 AI 交互排版口令")
    user_cmd = st.text_area(
        "支持极度随性的中文：",
        value="把第一行表头背景设为高端淡蓝色，加粗，字号换成16号，全表居中",
        height=100
    )

with col_right:
    st.subheader("🖥️ 3. 数据预览与 AI 运行面板")

    shp_target_path = None
    tmp_dir = "uploaded_temp"

    if uploaded_files:
        os.makedirs(tmp_dir, exist_ok=True)
        for f in uploaded_files:
            with open(os.path.join(tmp_dir, f.name), "wb") as buffer:
                buffer.write(f.read())
            if f.name.endswith(".shp"):
                shp_target_path = os.path.join(tmp_dir, f.name)

    if shp_target_path:
        try:
            gdf = gpd.read_file(shp_target_path)
            df_raw = pd.DataFrame(gdf.drop(columns='geometry', errors='ignore'))
            st.write("📊 原始属性表数据预览：")
            st.dataframe(df_raw.head(5), use_container_width=True)
        except Exception as e:
            st.error(f"读取SHP文件失败，请确保同时上传了.shp, .shx和.dbf文件！")
            shp_target_path = None
    else:
        st.info("💡 请在左侧上传您的 GIS 矢量数据集组合开始体验。")

    # 触发按钮
    if st.button("🚀 开始验证权限并执行 AI 智能排版", type="primary", use_container_width=True):
        if not shp_target_path:
            st.error("错误：请先上传完整的 SHP 文件组合！")
        elif not user_cmd.strip():
            st.error("错误：排版口令不能为空！")
        # 鉴权拦截
        elif not st.session_state.logged_in and not st.session_state.pay_unlocked:
            st.error("🚫 权限验证拦截：请在左侧登录QQ邮箱会员，或者通过微信扫码单次付费解锁！")
        else:
            with st.spinner("🔒 权限已解锁！AI 正在重构您的 Excel..."):
                try:
                    out_excel = "web_styled_result.xlsx"
                    df_raw.to_excel(out_excel, index=False)

                    ai_json = ask_ai_to_parse_instruction(user_cmd)
                    st.info("💡 **AI 智能理解大脑输出：**")
                    st.json(ai_json)

                    apply_styles_to_excel(out_excel, ai_json)
                    st.success("🎉 **AI 排版重构成功！结果已就绪。**")

                    with open(out_excel, "rb") as file:
                        st.download_button(
                            label="📥 立即点击下载并弹出已排版的属性表 Excel",
                            data=file,
                            file_name=f"AI排版_结果表.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    if st.session_state.pay_unlocked:
                        st.session_state.pay_unlocked = False

                except Exception as ex:
                    st.error(f"系统运行崩溃: {ex}")